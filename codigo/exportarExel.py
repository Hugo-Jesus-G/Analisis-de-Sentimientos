from openpyxl import Workbook
import pandas as pd
class exportar:
    
    
    
    
    def exportarAExel(listaLimpia,coincidencias,origenCoincidencias,ruta,frecuenciaPositiva,frecuenciaNegativa ):

     

        wb = Workbook()
        hoja = wb.active

        hoja.append(["Comentario", "Coincidencia", "Origen","", "Negtivo", "frecuencia","","Positivo","frecuencia"])

        for comentario, coincidencia, origen in zip(listaLimpia, coincidencias, origenCoincidencias):
            hoja.append([comentario, ", ".join(coincidencia), ", ".join(origen)])
            
        for i, (clave, valor) in enumerate(frecuenciaPositiva.items(), start=2):  # Empezamos desde la segunda fila
            hoja.cell(row=i, column=5).value = clave  # Columna 4 para la clave
            hoja.cell(row=i, column=6).value = valor  # Columna 5 para el valor
            

        for i, (clave, valor) in enumerate(frecuenciaNegativa.items(), start=2):  # Empezamos desde la segunda fila
            hoja.cell(row=i, column=8).value = clave  # Columna 4 para la clave
            hoja.cell(row=i, column=9).value = valor  # Columna 5 para el valor
        
        wb.save(ruta)

    
    
    
        
        


        
        
        
        

        
        
        
        
        